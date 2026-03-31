from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from jiramodule.services.jira_client import JiraManager
from team.models import TeamDev


class SprintWorkbookBuilder:
    start_col = 3
    title_col = 2
    available_row = 7

    def __init__(self, jira: JiraManager, sprint_id: int) -> None:
        self.jira = jira
        self.sprint_id = sprint_id
        self.workbook = Workbook()
        self.fill_sum = PatternFill(start_color="ff6565", end_color="ff6565", fill_type="solid")
        self.fill_title = PatternFill(start_color="ff8f8f", end_color="ff8f8f", fill_type="solid")
        self.fill = PatternFill(start_color="ffb9b9", end_color="ffb9b9", fill_type="solid")
        self.teamdevs = list(
            TeamDev.objects.filter(team=jira.cfg.team).select_related("dev").order_by("order", "id")
        )
        self.member_names = [teamdev.dev.name for teamdev in self.teamdevs]
        self.state_col = self.start_col + len(self.member_names)
        self.sum_rows: list[int] = []
        self.done_statuses = self._build_done_statuses()
        self.current_sprint_id = sprint_id
        self.next_sprint_id = (self.jira.get_next_sprint(self.jira.cfg.jira_board_id) or {}).get("id")
        self.project_id = self.jira.cfg.jira_project_id
        self.recurrent_epic_key = self.jira.cfg.jira_recurrent_epic_key

    def build(self) -> bytes:
        capacity_sheet = self.workbook.active
        capacity_sheet.title = "Capacity sheet"
        self._populate_capacity_sheet(capacity_sheet)

        reality_sheet = self.workbook.create_sheet("Estimation VS Reality")
        self._populate_reality_sheet(reality_sheet)

        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _build_done_statuses(self) -> str:
        status_groups = self.jira.get_dict_statuses()
        done_statuses = status_groups.get("Done", {}).get("statuses", [])
        return ",".join(f'"{status["name"]}"' for status in done_statuses)

    def _populate_capacity_sheet(self, worksheet) -> None:
        row = self._write_header(worksheet, 1)
        row = self._write_section(worksheet, "RECURRENTS", row + 1, [self._jql_recurrents()])
        row = self._write_section(
            worksheet,
            "PROJECTS",
            row + 1,
            [self._jql_projects_current(), self._jql_projects_next()],
        )
        row = self._write_section(worksheet, "BKL", row + 1, [self._jql_backlog_remaining()])
        row = self._write_section(worksheet, "NEW BKL", row + 1, [self._jql_new_backlog()])
        row = self._write_total_sum(worksheet, row)
        self._apply_sheet_border(worksheet, row - 1)

    def _populate_reality_sheet(self, worksheet) -> None:
        row = self._write_header(worksheet, 1)
        row = self._write_section(worksheet, "BKL", row, [self._jql_backlog_done()], diff_spent_hours=True)
        self._write_section(
            worksheet,
            "PROJECTS",
            row,
            [self._jql_non_backlog_done()],
            diff_spent_hours=True,
        )

    def _write_header(self, worksheet, row: int) -> int:
        for column in range(1, self.start_col):
            worksheet.cell(row=row, column=column).fill = self.fill_sum

        for index, name in enumerate(self.member_names):
            cell = worksheet.cell(row=row, column=self.start_col + index, value=name)
            cell.fill = self.fill_sum

        cell = worksheet.cell(row=row, column=self.state_col, value="State")
        cell.fill = self.fill_sum
        worksheet.freeze_panes = "A2"
        row += 1

        row = self._write_member_metric_row(worksheet, row, "Capacity", lambda teamdev: teamdev.dev.workload)
        row = self._write_member_metric_row(worksheet, row, "Holidays", lambda teamdev: teamdev.dev.holidays)
        row = self._write_member_metric_row(worksheet, row, "Fastlane (10%)", lambda teamdev: teamdev.dev.fastlane)
        row = self._write_member_metric_row(
            worksheet,
            row,
            "Project meetings",
            lambda teamdev: teamdev.dev.project_meetings,
        )
        row = self._write_member_metric_row(worksheet, row, "Meetings", lambda teamdev: teamdev.dev.other_meetings)
        return self._write_available_row(worksheet, row)

    def _write_member_metric_row(self, worksheet, row: int, label: str, getter) -> int:
        worksheet.cell(row=row, column=self.start_col - 1, value=label)
        for index, teamdev in enumerate(self.teamdevs):
            cell = worksheet.cell(row=row, column=self.start_col + index, value=getter(teamdev))
            cell.fill = self.fill
        return row + 1

    def _write_available_row(self, worksheet, row: int) -> int:
        worksheet.cell(row=row, column=self.start_col - 2).fill = self.fill_sum
        worksheet.cell(row=row, column=self.start_col - 1, value="Available").fill = self.fill_sum

        for index in range(len(self.member_names)):
            column = self.start_col + index
            column_letter = get_column_letter(column)
            formula = (
                f"={column_letter}{row-5}-{column_letter}{row-4}-{column_letter}{row-3}-"
                f"{column_letter}{row-2}-{column_letter}{row-1}"
            )
            cell = worksheet.cell(row=row, column=column, value=formula)
            cell.fill = self.fill_sum
        return row + 1

    def _write_section(self, worksheet, title: str, row: int, jqls: list[str], diff_spent_hours: bool = False) -> int:
        worksheet.cell(row=row, column=1, value=title)
        for column in range(1, 3 + len(self.member_names)):
            worksheet.cell(row=row, column=column).fill = self.fill_title

        row += 1
        first_row = row
        for jql in jqls:
            for issue in self.jira.run_jql(jql):
                row = self._write_issue_row(worksheet, row, issue, diff_spent_hours)

        return self._write_sum_row(worksheet, row, first_row)

    def _write_issue_row(self, worksheet, row: int, issue: dict, diff_spent_hours: bool) -> int:
        first_name = self._get_firstname(issue)
        if first_name == "Renaud":
            return row

        worksheet.cell(row=row, column=self.title_col, value=self._issue_title(issue))
        if first_name and first_name in self.member_names:
            column = self.start_col + self.member_names.index(first_name)
            worksheet.cell(
                row=row,
                column=column,
                value=self._resolve_issue_workload(issue, diff_spent_hours),
            )

        worksheet.cell(row=row, column=self.state_col, value=issue.get("state"))
        return row + 1

    def _write_sum_row(self, worksheet, row: int, first_row: int) -> int:
        worksheet.cell(row=row, column=1, value="Sum").fill = self.fill_sum
        for column in range(1, self.start_col):
            worksheet.cell(row=row, column=column).fill = self.fill_sum

        for index in range(len(self.member_names)):
            column = self.start_col + index
            column_letter = get_column_letter(column)
            cell = worksheet.cell(row=row, column=column)
            cell.fill = self.fill_sum
            cell.value = f"=SUM({column_letter}{first_row}:{column_letter}{row - 1})" if row - 1 >= first_row else 0

        self.sum_rows.append(row)
        return row + 1

    def _write_total_sum(self, worksheet, row: int) -> int:
        worksheet.cell(row=row, column=1, value="TOTAL")
        for column in range(1, self.start_col):
            worksheet.cell(row=row, column=column).fill = self.fill_title

        for index in range(len(self.member_names)):
            column = self.start_col + index
            column_letter = get_column_letter(column)
            total_cell = worksheet.cell(row=row, column=column)
            total_cell.fill = self.fill_title
            total_cell.value = f"=SUM({','.join(f'{column_letter}{sum_row}' for sum_row in self.sum_rows)})" if self.sum_rows else 0

            delta_cell = worksheet.cell(
                row=row + 1,
                column=column,
                value=f"={column_letter}{self.available_row}-{column_letter}{row}",
            )
            delta_cell.fill = self.fill_sum

        worksheet.cell(row=row + 1, column=1, value="DELTA").fill = self.fill_sum
        worksheet.cell(row=row + 1, column=2, value="").fill = self.fill_sum
        return row + 2

    def _apply_sheet_border(self, worksheet, end_row: int) -> None:
        border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )
        end_column = max(self.state_col, 10)
        for row_cells in worksheet.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=end_column):
            for cell in row_cells:
                cell.border = border

    def _get_firstname(self, issue: dict) -> str | None:
        assignee = issue.get("assignee") or {}
        display_name = assignee.get("displayName")
        if not display_name:
            return None
        parts = display_name.split(", ")
        return parts[1] if len(parts) > 1 else display_name

    def _issue_title(self, issue: dict) -> str:
        estimate = issue.get("estimation", {}).get("original_estimate_hours") or 0
        return f"{issue.get('key', '')} - {issue.get('summary', '')} ({estimate})"

    def _resolve_issue_workload(self, issue: dict, diff_spent_hours: bool) -> float:
        if diff_spent_hours:
            estimated_hours = float(issue.get("estimation", {}).get("original_estimate_hours") or 0)
            spent_hours = float(issue.get("spent_hours") or 0)
            return estimated_hours - spent_hours
        return float(issue.get("remaining_hours") or 0)

    def _jql_recurrents(self) -> str:
        return (
            f"project = {self.project_id} and sprint = {self.next_sprint_id} "
            f"and 'Epic Link' = {self.recurrent_epic_key}"
        )

    def _jql_projects_current(self) -> str:
        return (
            f"project != {self.project_id} and sprint = {self.current_sprint_id} "
            f"and status not in ({self.done_statuses}) and "
            f"(labels = \"ASSU_POPU_SPRINT\" or issuetype in standardIssueTypes())"
        )

    def _jql_projects_next(self) -> str:
        return (
            f"project != {self.project_id} and sprint = {self.next_sprint_id} and "
            f"(labels = \"ASSU_POPU_SPRINT\" or issuetype in standardIssueTypes())"
        )

    def _jql_backlog_remaining(self) -> str:
        return (
            f"project = {self.project_id} and sprint = {self.current_sprint_id} and "
            f"status not in ({self.done_statuses}) and {self.jira.JIRA_ONLY_STANDARD_TYPES} ORDER BY status ASC"
        )

    def _jql_new_backlog(self) -> str:
        return (
            f"project = {self.project_id} and sprint = {self.next_sprint_id} and "
            f"('Epic Link' != {self.recurrent_epic_key} OR 'Epic Link' is EMPTY)"
        )

    def _jql_backlog_done(self) -> str:
        return (
            f"project = {self.project_id} and sprint = {self.current_sprint_id} and "
            f"status in ({self.done_statuses}) and {self.jira.JIRA_ONLY_STANDARD_TYPES} ORDER BY status ASC"
        )

    def _jql_non_backlog_done(self) -> str:
        return (
            f"project != {self.project_id} and sprint = {self.current_sprint_id} and "
            f"status in ({self.done_statuses}) and {self.jira.JIRA_ONLY_STANDARD_TYPES} ORDER BY status ASC"
        )


def create_xlsx_bytes(jm: JiraManager, sprint_id: int) -> bytes:
    return SprintWorkbookBuilder(jm, sprint_id).build()
