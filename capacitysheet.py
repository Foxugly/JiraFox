from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from data import FIRSTNAMES, JIRA_ONLY_STANDARD_TYPES, REC_EPIC_KEY, BOARD_ID, PROJECT_ID, dict_statuses
from jiramanager import JiraManager


def create_xlsx(path: str, jm: "JiraManager"):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Capacity sheet"

    start_col = 3  # C
    title_col = 2  # B
    state_col = start_col + len(FIRSTNAMES)  # colonne après les prénoms

    ordered_FIRSTNAMES = sorted(FIRSTNAMES.items(), key=lambda x: x[1])

    list_sum_rows: list[int] = []

    def write_header(ws, row: int) -> int:
        for i, (name, _) in enumerate(ordered_FIRSTNAMES):
            ws.cell(row=row, column=start_col + i, value=name)
        ws.cell(row=row, column=state_col, value="State")
        return row + 1

    def get_firstname(issue: dict) -> str | None:
        assignee = issue.get("assignee")
        if not assignee or not assignee.get("displayName"):
            return None
        parts = assignee["displayName"].split(", ")
        return parts[1] if len(parts) > 1 else assignee["displayName"]

    def issue_title(issue: dict) -> str:
        key = issue.get("key", "")
        summary = issue.get("summary", "")
        est = issue.get("estimation", {}).get("original_estimate_hours")
        est_txt = "" if est is None else str(est)
        return f"{key} - {summary} ({est_txt})"

    def write_issue_row(ws, row: int, issue: dict, diff_spent_hours: bool) -> int:
        firstname = get_firstname(issue)

        if firstname == "Renaud":
            return row

        ws.cell(row=row, column=title_col, value=issue_title(issue))

        if firstname and firstname in FIRSTNAMES:
            col = start_col + FIRSTNAMES[firstname]
            if diff_spent_hours:
                estimated_hours = float(issue.get("estimation").get("original_estimate_hours"))
                spent_hours = float(issue.get("spent_hours"))
                value = estimated_hours - spent_hours
            else:
                value = float(issue.get("remaining_hours", 0))
            ws.cell(row=row, column=col, value=value)

        ws.cell(row=row, column=state_col, value=issue.get("state"))
        return row + 1

    def write_sum_row(ws, row: int, first_row: int) -> int:
        # Somme par colonne (prénoms)
        for i in range(len(FIRSTNAMES)):
            col = start_col + i
            col_letter = get_column_letter(col)

            if row - 1 >= first_row:
                ws.cell(row=row, column=col, value=f"=SUM({col_letter}{first_row}:{col_letter}{row - 1})")
            else:
                ws.cell(row=row, column=col, value=0)

        list_sum_rows.append(row)
        return row + 1

    def write_section(ws, title: str, row: int, jqls: list[str], diff_spent_hours=False) -> int:
        ws.cell(row=row, column=1, value=title)
        row += 1
        first_row = row

        for jql in jqls:
            for issue in jm.run_jql(jql):
                row = write_issue_row(ws, row, issue, diff_spent_hours)

        # Ligne SUM
        row = write_sum_row(ws, row, first_row)
        return row

    def write_total_sum(ws, row: int) -> int:
        ws.cell(row=row, column=1, value="TOTAL")

        # Total par colonne = somme des sommes intermédiaires
        for i in range(len(FIRSTNAMES)):
            col = start_col + i
            col_letter = get_column_letter(col)

            if not list_sum_rows:
                ws.cell(row=row, column=col, value=0)
            else:
                # ex: =SUM(C5,C12,C20)
                refs = ",".join([f"{col_letter}{r}" for r in list_sum_rows])
                ws.cell(row=row, column=col, value=f"=SUM({refs})")

        return row + 1

    # --- Header
    row = 1
    row = write_header(ws1, row)

    done_statuses = ",".join([f'"{s["name"]}"' for s in dict_statuses["Done"]["statuses"]])
    current_sprint = jm.get_current_sprint(BOARD_ID)
    current_sprint_id = current_sprint.get("id")
    next_sprint = jm.get_next_sprint(BOARD_ID)
    next_sprint_id = next_sprint.get("id")
    # --- Sections
    jql_bkl = (
        f"project = {PROJECT_ID} and sprint = {current_sprint_id} "
        f"and status not in ({done_statuses}) and {JIRA_ONLY_STANDARD_TYPES} ORDER BY status ASC"
    )
    row = write_section(ws1, "BKL", row, [jql_bkl], False)

    jql_recurrents = f"project = {PROJECT_ID} and sprint = {next_sprint_id} and 'Epic Link' = {REC_EPIC_KEY}"
    row = write_section(ws1, "RECURRENTS", row, [jql_recurrents], False)

    jql_new_bkl = f"project = {PROJECT_ID} and sprint = {next_sprint_id} and 'Epic Link' != {REC_EPIC_KEY}"
    row = write_section(ws1, "NEW BKL", row, [jql_new_bkl], False)

    jql_projects_current = (
        f"project != {PROJECT_ID} and sprint = {current_sprint_id} and status not in ({done_statuses})"
    )
    jql_projects_next = f"project != {PROJECT_ID} and sprint = {next_sprint_id}"
    row = write_section(ws1, "PROJECTS", row, [jql_projects_current, jql_projects_next], False)

    # --- TOTAL
    row = write_total_sum(ws1, row)

    ws2 = wb.create_sheet('Estimation VS Reality')
    # --- Header
    row = 1
    row = write_header(ws2, row)
    jql_bkl = (
        f"project = {PROJECT_ID} and sprint = {current_sprint_id} "
        f"and status in ({done_statuses}) and {JIRA_ONLY_STANDARD_TYPES} ORDER BY status ASC"
    )
    row = write_section(ws2, "BKL", row, [jql_bkl], True)
    jql_not_bkl = (
        f"project != {PROJECT_ID} and sprint = {current_sprint_id} "
        f"and status in ({done_statuses}) and {JIRA_ONLY_STANDARD_TYPES} ORDER BY status ASC"
    )
    row = write_section(ws2, "PROJECTS", row, [jql_not_bkl], True)

    wb.save(path)
